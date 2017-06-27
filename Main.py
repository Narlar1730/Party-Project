import time
from random import randint
import urllib.request



import PyPDF2
from openpyxl import load_workbook








# Initializing all the stuff that I want to find
NetSales = 0.0
GrossSales = 0.0
Customers = 0.0
VisaTotal = 0.0
MasterCard = 0.0
Amex = 0.0
DebitTotal = 0.0
TotalCash = 0.0
Discounts = 0.0
Voids = 0.0
Refunds = 0.0
Outgoings = 0.0

# Opens the file to read
pdfFileobj = open('test.pdf', 'rb')

# Reads the file in
pdfReader = PyPDF2.PdfFileReader(pdfFileobj)

# This reads the two files at the same time, and then removes some redundant infromation
# primarily getting rid of the commas, because it just makes sense to get rid of them now rather than later

page1 = pdfReader.getPage(0)
page2 = pdfReader.getPage(1)
page1txt = page1.extractText()
page1txt = page1txt.replace(',', '')
page2txt = page2.extractText()
page2txt = page2txt.replace(',', '')

# Extracts the net sales from the pdf
Netstr = page1txt.split("Net Sales$", 1)[1]
Netstr = Netstr.split("--", 1)[0]
NetSales = float(Netstr)

#Extracts the gross sales and the customer count from the report
GrossStr = page1txt.split("Gross Sales$", 1)[1]
cstmrStr = GrossStr.split("%", 1)[1]
cstmrStr = cstmrStr.split("Q", 1)[0]
GrossStr = GrossStr.split("$", 1)[0]
GrossSales = float(GrossStr)
Customers = float(cstmrStr)

# Here we attempt to extract the visa total
VisaStr = page1txt.split("Visa Total", 1)[1]
VisaStr = VisaStr.split("A", 1)[0]
VisaStr = VisaStr.split("$", 1)[1]
VisaTotal = float(VisaStr)

# Here we extract the master card total
MstrStr = page1txt.split("MasterCard Total", 1)[1]
MstrStr = MstrStr.split("S", 1)[0]
MstrStr = MstrStr.split("$", 1)[1]
MasterCard = float(MstrStr)

# Here we extract the amex total
AmxStr = page1txt.split("Amex Total", 1)[1]
AmxStr = AmxStr.split("T", 1)[0]
AmxStr = AmxStr.split("$", 1)[1]
Amex = float(AmxStr)

# Here we extract the total cash payments
CshStr = page1txt.split("Total Cash Payments$", 1)[1]
CshStr = CshStr.split("Cash", 1)[0]
TotalCash = float(CshStr)

# Here we extract the total debit payments
DbtStr = page1txt.split("Total Debit Payments$", 1)[1]
DbtStr = DbtStr.split("M", 1)[0]
DebitTotal = float(DbtStr)

# Here we extract the total outgoings.
try:
    OutStr = page1txt.split("SummaryQtyAmountQtyAmount$", 1)[1]
    OutStr = OutStr.split("O", 1)[0]
    Outgoings = float(OutStr)
except IndexError:
    print ("No Outgoings")

# Here we extract the total Voids
VoidStr = page2txt.split("Voids Total", 1)[1]
VoidStr = VoidStr.split("Re", 1)[0]
VoidStr = VoidStr.split("$", 1)[1]
Voids = float(VoidStr)

# Finally Here we extract the discounts
DsctStr = page2txt.split("Total Discounts", 1)[1]
DsctStr = DsctStr.split("Re", 1)[0]
DsctStr = DsctStr.split("$", 1)[1]
Discounts = float(DsctStr)

# Here we get the information from the internet using HTML scraping. We have to
# connect to the right page


# Here we write all of the data to the correct column
# First step is loading and creating the worksheet
wb = load_workbook('Schnitz EOD WE 11.06.17.xlsx')

names = wb.sheetnames
ws = wb.active
currentDate = time.strftime("%y-%m-%d")
CellDate = ""
ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
Count= 1

# Here we idnetify the correct cells to write to, by matching to the current date.
for i in range(0, 7):
    pos = ALPHABET[Count]
    CellDate = str(ws[pos+'12'].value)
    CellDate = CellDate.split(" ", 1)[0]
    CellDate = CellDate.split("20", 1)[1]
    if CellDate == currentDate:
        print "SUCCESS", CellDate
        break
    Count += 1

ColumnToWrite = ALPHABET[Count]
# Here we write to the desired column, what we want to write
ws[ColumnToWrite+'13'] = TotalCash
ws[ColumnToWrite+'14'] = DebitTotal
ws[ColumnToWrite+'15'] = MasterCard
ws[ColumnToWrite+'16'] = Amex
ws[ColumnToWrite+'17'] = VisaTotal
ws[ColumnToWrite+'18'] = 0
ws[ColumnToWrite+'20'] = Outgoings
ws[ColumnToWrite+'24'] = Customers
ws[ColumnToWrite+'40'] = NetSales
ws[ColumnToWrite+'58'] = 0
ws[ColumnToWrite+'59'] = 0
ws[ColumnToWrite+'60'] = Discounts
ws[ColumnToWrite+'61'] = Voids
ws[ColumnToWrite+'83'] = "None"
ws[ColumnToWrite+'84'] = "None"
# This does some nonsense and is really funny, putting randomly generated notes at the bottom coz cbf.
NightGross = ws[ColumnToWrite+'38']
Comment = ""
quietNightComments = ["cleaning", "training new skills", "cleaning back of house", "improving customer interactions"]
avNightComments = ["to improve customer interactions", "to do extra cleaning out the front", "to sweep outside areas", "to improve wait times"]
if NightGross < 600:
    cmt = quietNightComments[randint[0, 3]]
    Comment = "Quiet night, focused on " + cmt
elif NightGross < 800:
    cmt = avNightComments[randint[0, 3]]
    Comment = "Average night, tried " + cmt
else:
    Comment = "Busy night"

ws[ColumnToWrite+'85'] = Comment
wb.save('Schnitz EOD WE 11.06.17.xlsx')

# This prints out our results, it will be replaced later
print "Gross Sales: ", GrossSales, "\nTotal Customers: ", Customers, "\nNet sales: ", NetSales, "\nVisa Total: ", VisaTotal, "\nMasterCard Total: ", MasterCard, "\nAmex Total: ", Amex, "\nCash Total: ", TotalCash, "\nDebit Total:", DebitTotal, "\nOutgoings Total: ", Outgoings, "\nVoids: ", Voids, "\nDiscounts: ", Discounts
